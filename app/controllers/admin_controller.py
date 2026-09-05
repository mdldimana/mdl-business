import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
from flask_login import login_required, current_user
from app import db, bcrypt
from app.models import Utilisateur, Produit, Categorie, Commande
from datetime import datetime, timedelta
from app.utils import upload_image

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


# ============================================
# DÉCORATEUR POUR VÉRIFIER LES DROITS ADMIN
# ============================================
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.est_admin:
            flash('Accès non autorisé. Vous devez être administrateur.', 'danger')
            return redirect(url_for('accueil'))
        return f(*args, **kwargs)

    return decorated_function


# ============================================
# TABLEAU DE BORD
# ============================================
@admin_bp.route('/')
@login_required
@admin_required
def tableau_bord():
    """Tableau de bord de l'administration"""

    # Statistiques
    total_commandes = Commande.query.count()
    total_produits = Produit.query.count()
    total_utilisateurs = Utilisateur.query.count()

    # Chiffre d'affaires
    commandes_payees = Commande.query.filter(Commande.statut.in_(['payee', 'expediee', 'livree'])).all()
    ca_total = sum(c.total for c in commandes_payees)

    # Commandes récentes
    commandes_recentes = Commande.query.order_by(Commande.date_creation.desc()).limit(10).all()

    # Produits en rupture de stock
    produits_rupture = Produit.query.filter(Produit.stock <= 0).count()
    produits_stock_bas = Produit.query.filter(Produit.stock.between(1, 5)).count()

    # Commandes par statut
    commandes_par_statut = {
        'en_attente': Commande.query.filter_by(statut='en_attente').count(),
        'payee': Commande.query.filter_by(statut='payee').count(),
        'expediee': Commande.query.filter_by(statut='expediee').count(),
        'livree': Commande.query.filter_by(statut='livree').count(),
        'annulee': Commande.query.filter_by(statut='annulee').count(),
    }

    return render_template('admin/tableau_bord.html',
                           total_commandes=total_commandes,
                           total_produits=total_produits,
                           total_utilisateurs=total_utilisateurs,
                           ca_total=ca_total,
                           commandes_recentes=commandes_recentes,
                           produits_rupture=produits_rupture,
                           produits_stock_bas=produits_stock_bas,
                           commandes_par_statut=commandes_par_statut)


# ============================================
# GESTION DES PRODUITS
# ============================================
@admin_bp.route('/produits')
@login_required
@admin_required
def gestion_produits():
    """Liste des produits pour l'administration"""
    produits = Produit.query.all()
    categories = Categorie.query.all()
    return render_template('admin/produits.html', produits=produits, categories=categories)


@admin_bp.route('/produit/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_produit():
    """Ajoute un nouveau produit"""
    if request.method == 'POST':
        nom = request.form.get('nom')
        slug = request.form.get('slug')
        description = request.form.get('description')
        prix = float(request.form.get('prix', 0))
        prix_promo = request.form.get('prix_promo')
        stock = int(request.form.get('stock', 0))
        categorie_id = int(request.form.get('categorie_id', 0))
        est_disponible = request.form.get('est_disponible') == 'on'

        if not nom or not slug or prix <= 0:
            flash('Veuillez remplir tous les champs obligatoires.', 'danger')
            return redirect(url_for('admin.ajouter_produit'))

        # Gérer l'upload de l'image vers Cloudinary
        image_path = None
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename != '':
                try:
                    image_path = upload_image(file)
                    if not image_path:
                        flash('Erreur lors de l\'upload de l\'image.', 'danger')
                        return redirect(url_for('admin.ajouter_produit'))
                except ValueError as e:
                    flash(str(e), 'danger')
                    return redirect(url_for('admin.ajouter_produit'))

        produit = Produit(
            nom=nom,
            slug=slug,
            description=description,
            prix=prix,
            prix_promo=float(prix_promo) if prix_promo else None,
            stock=stock,
            categorie_id=categorie_id,
            est_disponible=est_disponible,
            image_principale=image_path
        )

        db.session.add(produit)
        db.session.commit()
        flash(f'Produit "{nom}" ajouté avec succès !', 'success')
        return redirect(url_for('admin.gestion_produits'))

    categories = Categorie.query.all()
    return render_template('admin/ajouter_produit.html', categories=categories)


@admin_bp.route('/produit/modifier/<int:produit_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_produit(produit_id):
    """Modifie un produit existant"""
    produit = Produit.query.get_or_404(produit_id)

    if request.method == 'POST':
        produit.nom = request.form.get('nom')
        produit.slug = request.form.get('slug')
        produit.description = request.form.get('description')
        produit.prix = float(request.form.get('prix', 0))
        prix_promo = request.form.get('prix_promo')
        produit.prix_promo = float(prix_promo) if prix_promo else None
        produit.stock = int(request.form.get('stock', 0))
        produit.categorie_id = int(request.form.get('categorie_id', 0))
        produit.est_disponible = produit.stock > 0

        # Gérer l'upload de l'image vers Cloudinary
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename != '':
                try:
                    image_url = upload_image(file)
                    if image_url:
                        produit.image_principale = image_url
                    else:
                        flash('Erreur lors de l\'upload de l\'image.', 'danger')
                        return redirect(url_for('admin.modifier_produit', produit_id=produit_id))
                except ValueError as e:
                    flash(str(e), 'danger')
                    return redirect(url_for('admin.modifier_produit', produit_id=produit_id))

        db.session.commit()
        flash(f'Produit "{produit.nom}" modifié avec succès !', 'success')
        return redirect(url_for('admin.gestion_produits'))

    categories = Categorie.query.all()
    return render_template('admin/modifier_produit.html', produit=produit, categories=categories)


@admin_bp.route('/produit/supprimer/<int:produit_id>', methods=['POST'])
@login_required
@admin_required
def supprimer_produit(produit_id):
    """Supprime un produit"""
    produit = Produit.query.get_or_404(produit_id)
    nom = produit.nom
    db.session.delete(produit)
    db.session.commit()
    flash(f'Produit "{nom}" supprimé.', 'success')
    return redirect(url_for('admin.gestion_produits'))


# ============================================
# GESTION DES COMMANDES
# ============================================
@admin_bp.route('/commandes')
@login_required
@admin_required
def gestion_commandes():
    """Liste des commandes pour l'administration"""
    statut_filtre = request.args.get('statut', '')

    query = Commande.query
    if statut_filtre:
        query = query.filter_by(statut=statut_filtre)

    commandes = query.order_by(Commande.date_creation.desc()).all()
    statuts = ['en_attente', 'payee', 'expediee', 'livree', 'annulee']

    return render_template('admin/commandes.html',
                           commandes=commandes,
                           statuts=statuts,
                           statut_filtre=statut_filtre)


@admin_bp.route('/commande/modifier-statut/<int:commande_id>', methods=['POST'])
@login_required
@admin_required
def modifier_statut(commande_id):
    """Modifie le statut d'une commande"""
    commande = Commande.query.get_or_404(commande_id)
    nouveau_statut = request.form.get('statut')

    if nouveau_statut in ['en_attente', 'payee', 'expediee', 'livree', 'annulee']:
        commande.statut = nouveau_statut
        db.session.commit()
        flash(f'Statut de la commande #{commande.reference} mis à jour.', 'success')

    return redirect(url_for('admin.gestion_commandes'))


# ============================================
# GESTION DES UTILISATEURS
# ============================================
@admin_bp.route('/utilisateurs')
@login_required
@admin_required
def gestion_utilisateurs():
    """Liste des utilisateurs"""
    utilisateurs = Utilisateur.query.all()
    return render_template('admin/utilisateurs.html', utilisateurs=utilisateurs)


@admin_bp.route('/utilisateur/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_utilisateur():
    """Ajoute un nouvel utilisateur"""
    if request.method == 'POST':
        email = request.form.get('email')
        prenom = request.form.get('prenom')
        nom = request.form.get('nom')
        telephone = request.form.get('telephone')
        mot_de_passe = request.form.get('mot_de_passe')
        role = request.form.get('role', 'client')
        est_actif = request.form.get('est_actif') == 'on'

        if Utilisateur.query.filter_by(email=email).first():
            flash('Cet email est déjà utilisé.', 'danger')
            return redirect(url_for('admin.ajouter_utilisateur'))

        if not all([email, prenom, nom, mot_de_passe]):
            flash('Veuillez remplir tous les champs obligatoires.', 'danger')
            return redirect(url_for('admin.ajouter_utilisateur'))

        mot_de_passe_hash = bcrypt.generate_password_hash(mot_de_passe).decode('utf-8')

        utilisateur = Utilisateur(
            email=email,
            prenom=prenom,
            nom=nom,
            mot_de_passe_hash=mot_de_passe_hash,
            role=role,
            est_admin=(role == 'admin'),
            est_actif=est_actif
        )

        db.session.add(utilisateur)
        db.session.commit()

        flash(f'Utilisateur "{prenom} {nom}" créé avec succès ! Rôle : {role}', 'success')
        return redirect(url_for('admin.gestion_utilisateurs'))

    return render_template('admin/ajouter_utilisateur.html')


@admin_bp.route('/utilisateur/modifier/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_utilisateur(user_id):
    """Modifie un utilisateur existant"""
    utilisateur = Utilisateur.query.get_or_404(user_id)

    if utilisateur.id == current_user.id:
        flash('Vous ne pouvez pas modifier votre propre compte.', 'warning')
        return redirect(url_for('admin.gestion_utilisateurs'))

    if request.method == 'POST':
        utilisateur.prenom = request.form.get('prenom')
        utilisateur.nom = request.form.get('nom')
        utilisateur.email = request.form.get('email')
        utilisateur.telephone = request.form.get('telephone')
        utilisateur.role = request.form.get('role', 'client')
        utilisateur.est_admin = (utilisateur.role == 'admin')
        utilisateur.est_actif = request.form.get('est_actif') == 'on'

        nouveau_mdp = request.form.get('mot_de_passe')
        if nouveau_mdp and len(nouveau_mdp) >= 6:
            utilisateur.mot_de_passe_hash = bcrypt.generate_password_hash(nouveau_mdp).decode('utf-8')
            flash('Le mot de passe a été modifié.', 'success')

        db.session.commit()
        flash(f'Utilisateur "{utilisateur.prenom} {utilisateur.nom}" modifié avec succès !', 'success')
        return redirect(url_for('admin.gestion_utilisateurs'))

    return render_template('admin/modifier_utilisateur.html', utilisateur=utilisateur)


@admin_bp.route('/utilisateur/supprimer/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def supprimer_utilisateur(user_id):
    """Supprime un utilisateur"""
    utilisateur = Utilisateur.query.get_or_404(user_id)

    if utilisateur.id == current_user.id:
        flash('Vous ne pouvez pas supprimer votre propre compte.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))

    admins = Utilisateur.query.filter_by(est_admin=True).count()
    if utilisateur.est_admin and admins <= 1:
        flash('Impossible de supprimer le dernier administrateur.', 'danger')
        return redirect(url_for('admin.gestion_utilisateurs'))

    nom = utilisateur.prenom + ' ' + utilisateur.nom
    db.session.delete(utilisateur)
    db.session.commit()

    flash(f'Utilisateur "{nom}" supprimé avec succès.', 'success')
    return redirect(url_for('admin.gestion_utilisateurs'))


# ============================================
# GESTION DES CATÉGORIES
# ============================================
@admin_bp.route('/categories')
@login_required
@admin_required
def gestion_categories():
    """Liste des catégories"""
    categories = Categorie.query.order_by(Categorie.nom).all()
    return render_template('admin/categories.html', categories=categories)


@admin_bp.route('/categorie/ajouter', methods=['GET', 'POST'])
@login_required
@admin_required
def ajouter_categorie():
    """Ajoute une nouvelle catégorie"""
    if request.method == 'POST':
        nom = request.form.get('nom')
        slug = request.form.get('slug')
        description = request.form.get('description')
        est_active = request.form.get('est_active') == 'on'

        if not nom or not slug:
            flash('Veuillez remplir tous les champs obligatoires.', 'danger')
            return redirect(url_for('admin.ajouter_categorie'))

        if Categorie.query.filter_by(slug=slug).first():
            flash('Ce slug est déjà utilisé.', 'danger')
            return redirect(url_for('admin.ajouter_categorie'))

        categorie = Categorie(
            nom=nom,
            slug=slug,
            description=description,
            est_active=est_active
        )

        db.session.add(categorie)
        db.session.commit()

        flash(f'Catégorie "{nom}" ajoutée avec succès !', 'success')
        return redirect(url_for('admin.gestion_categories'))

    return render_template('admin/ajouter_categorie.html')


@admin_bp.route('/categorie/modifier/<int:categorie_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_categorie(categorie_id):
    """Modifie une catégorie existante"""
    categorie = Categorie.query.get_or_404(categorie_id)

    if request.method == 'POST':
        categorie.nom = request.form.get('nom')
        categorie.slug = request.form.get('slug')
        categorie.description = request.form.get('description')
        categorie.est_active = request.form.get('est_active') == 'on'

        existant = Categorie.query.filter(
            Categorie.slug == categorie.slug,
            Categorie.id != categorie.id
        ).first()
        if existant:
            flash('Ce slug est déjà utilisé par une autre catégorie.', 'danger')
            return redirect(url_for('admin.modifier_categorie', categorie_id=categorie_id))

        db.session.commit()
        flash(f'Catégorie "{categorie.nom}" modifiée avec succès !', 'success')
        return redirect(url_for('admin.gestion_categories'))

    return render_template('admin/modifier_categorie.html', categorie=categorie)


@admin_bp.route('/categorie/supprimer/<int:categorie_id>', methods=['POST'])
@login_required
@admin_required
def supprimer_categorie(categorie_id):
    """Supprime une catégorie"""
    categorie = Categorie.query.get_or_404(categorie_id)

    if categorie.produits:
        flash(f'Impossible de supprimer la catégorie "{categorie.nom}" car elle contient des produits.', 'danger')
        return redirect(url_for('admin.gestion_categories'))

    nom = categorie.nom
    db.session.delete(categorie)
    db.session.commit()

    flash(f'Catégorie "{nom}" supprimée avec succès.', 'success')
    return redirect(url_for('admin.gestion_categories'))


# ============================================
# EXPORTS PDF
# ============================================
@admin_bp.route('/export-produits-pdf')
@login_required
@admin_required
def export_produits_pdf():
    """Exporte la liste des produits en PDF avec logo et titre"""
    produits = Produit.query.all()

    from app.services.pdf_service import PDFService
    pdf = PDFService.generer_catalogue_produits(produits)

    response = current_app.response_class(pdf, mimetype='application/pdf')
    response.headers.set('Content-Disposition',
                         f'attachment; filename=catalogue_produits_{datetime.now().strftime("%Y%m%d")}.pdf')
    response.headers.set('Content-Type', 'application/pdf')

    return response


@admin_bp.route('/export-commandes-pdf')
@login_required
@admin_required
def export_commandes_pdf():
    """Exporte la liste des commandes en PDF avec logo et titre"""
    statut_filtre = request.args.get('statut', '')

    query = Commande.query
    if statut_filtre:
        query = query.filter_by(statut=statut_filtre)

    commandes = query.order_by(Commande.date_creation.desc()).all()

    from app.services.pdf_service import PDFService
    pdf = PDFService.generer_liste_commandes(commandes, statut_filtre)

    response = current_app.response_class(pdf, mimetype='application/pdf')
    response.headers.set('Content-Disposition',
                         f'attachment; filename=commandes_{datetime.now().strftime("%Y%m%d")}.pdf')
    response.headers.set('Content-Type', 'application/pdf')

    return response


@admin_bp.route('/export-utilisateurs-pdf')
@login_required
@admin_required
def export_utilisateurs_pdf():
    """Exporte la liste des utilisateurs en PDF avec logo et titre"""
    role_filtre = request.args.get('role', '')

    query = Utilisateur.query
    if role_filtre:
        query = query.filter_by(role=role_filtre)

    utilisateurs = query.all()

    from app.services.pdf_service import PDFService
    pdf = PDFService.generer_liste_utilisateurs(utilisateurs, role_filtre)

    response = current_app.response_class(pdf, mimetype='application/pdf')
    response.headers.set('Content-Disposition',
                         f'attachment; filename=utilisateurs_{datetime.now().strftime("%Y%m%d")}.pdf')
    response.headers.set('Content-Type', 'application/pdf')

    return response


@admin_bp.route('/export-categories-pdf')
@login_required
@admin_required
def export_categories_pdf():
    """Exporte la liste des catégories en PDF avec logo et titre"""
    categories = Categorie.query.all()

    from app.services.pdf_service import PDFService
    pdf = PDFService.generer_liste_categories(categories)

    response = current_app.response_class(pdf, mimetype='application/pdf')
    response.headers.set('Content-Disposition',
                         f'attachment; filename=categories_{datetime.now().strftime("%Y%m%d")}.pdf')
    response.headers.set('Content-Type', 'application/pdf')

    return response


# ============================================
# EXPORTS EXCEL (XLSX) - PROFESSIONNEL
# ============================================
@admin_bp.route('/export-produits-excel')
@login_required
@admin_required
def export_produits_excel():
    """Exporte la liste des produits en Excel (XLSX) avec titre, en-têtes et style"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    produits = Produit.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    # Styles
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill(start_color="0B2B4A", end_color="0B2B4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== TITRE =====
    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value="MDL BUSINESS - CATALOGUE DES PRODUITS")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Date
    ws.merge_cells('A2:I2')
    date_cell = ws.cell(row=2, column=1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ligne vide
    ws.row_dimensions[3].height = 10

    # ===== EN-TÊTES =====
    headers = ['ID', 'Nom', 'Slug', 'Description', 'Prix ($)', 'Prix Promo ($)', 'Stock', 'Catégorie', 'Disponible']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # ===== DONNÉES =====
    row_num = 5
    for p in produits:
        ws.cell(row=row_num, column=1, value=p.id)
        ws.cell(row=row_num, column=2, value=p.nom)
        ws.cell(row=row_num, column=3, value=p.slug)
        ws.cell(row=row_num, column=4, value=p.description or '')
        ws.cell(row=row_num, column=5, value=p.prix)
        ws.cell(row=row_num, column=6, value=p.prix_promo or '')
        ws.cell(row=row_num, column=7, value=p.stock)
        ws.cell(row=row_num, column=8, value=p.categorie.nom if p.categorie else '')
        ws.cell(row=row_num, column=9, value='Oui' if p.est_disponible else 'Non')

        for col in range(1, 10):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

        row_num += 1

    # Ajuster les largeurs
    for col in range(1, 10):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20

    # ===== PIED DE PAGE =====
    ws.merge_cells(f'A{row_num + 1}:I{row_num + 1}')
    footer_cell = ws.cell(row=row_num + 1, column=1, value="MDL Business - Tous droits réservés")
    footer_cell.font = Font(size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = current_app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers.set('Content-Disposition',
                         f'attachment; filename=produits_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return response


@admin_bp.route('/export-commandes-excel')
@login_required
@admin_required
def export_commandes_excel():
    """Exporte la liste des commandes en Excel (XLSX) avec titre, en-têtes et style"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    statut_filtre = request.args.get('statut', '')

    query = Commande.query
    if statut_filtre:
        query = query.filter_by(statut=statut_filtre)

    commandes = query.order_by(Commande.date_creation.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes"

    # Styles
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill(start_color="0B2B4A", end_color="0B2B4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value="MDL BUSINESS - LISTE DES COMMANDES")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sous-titre (filtre)
    ws.merge_cells('A2:G2')
    filtre_text = f"Filtre: {statut_filtre.replace('_', ' ').title()}" if statut_filtre else "Toutes les commandes"
    date_cell = ws.cell(row=2, column=1,
                        value=f"{filtre_text} - Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 10

    # En-têtes
    headers = ['#', 'Référence', 'Client', 'Email', 'Total ($)', 'Statut', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Données
    row_num = 5
    for i, c in enumerate(commandes, 1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=c.reference)
        ws.cell(row=row_num, column=3, value=f"{c.prenom} {c.nom}")
        ws.cell(row=row_num, column=4, value=c.email)
        ws.cell(row=row_num, column=5, value=float(c.total))
        ws.cell(row=row_num, column=6, value=c.statut.replace('_', ' ').title())
        ws.cell(row=row_num, column=7, value=c.date_creation.strftime('%d/%m/%Y %H:%M'))

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

        row_num += 1

    # Ajuster les largeurs
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20

    # Footer
    ws.merge_cells(f'A{row_num + 1}:G{row_num + 1}')
    footer_cell = ws.cell(row=row_num + 1, column=1, value="MDL Business - Tous droits réservés")
    footer_cell.font = Font(size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = current_app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers.set('Content-Disposition',
                         f'attachment; filename=commandes_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return response


@admin_bp.route('/export-utilisateurs-excel')
@login_required
@admin_required
def export_utilisateurs_excel():
    """Exporte la liste des utilisateurs en Excel (XLSX) avec titre, en-têtes et style"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    role_filtre = request.args.get('role', '')

    query = Utilisateur.query
    if role_filtre:
        query = query.filter_by(role=role_filtre)

    utilisateurs = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    # Styles
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill(start_color="0B2B4A", end_color="0B2B4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="MDL BUSINESS - LISTE DES UTILISATEURS")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sous-titre
    ws.merge_cells('A2:F2')
    filtre_text = f"Filtre: {role_filtre.title()}" if role_filtre else "Tous les utilisateurs"
    date_cell = ws.cell(row=2, column=1,
                        value=f"{filtre_text} - Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 10

    # En-têtes
    headers = ['ID', 'Email', 'Nom', 'Prénom', 'Rôle', 'Inscription', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Données
    row_num = 5
    for u in utilisateurs:
        ws.cell(row=row_num, column=1, value=u.id)
        ws.cell(row=row_num, column=2, value=u.email)
        ws.cell(row=row_num, column=3, value=u.nom or '')
        ws.cell(row=row_num, column=4, value=u.prenom or '')
        ws.cell(row=row_num, column=5, value=u.role.title() if u.role else 'Client')
        ws.cell(row=row_num, column=6, value=u.date_inscription.strftime('%d/%m/%Y') if u.date_inscription else '')
        ws.cell(row=row_num, column=7, value='Actif' if u.est_actif else 'Inactif')

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

        row_num += 1

    # Ajuster les largeurs
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20

    # Footer
    ws.merge_cells(f'A{row_num + 1}:F{row_num + 1}')
    footer_cell = ws.cell(row=row_num + 1, column=1, value="MDL Business - Tous droits réservés")
    footer_cell.font = Font(size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = current_app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers.set('Content-Disposition',
                         f'attachment; filename=utilisateurs_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return response


@admin_bp.route('/export-categories-excel')
@login_required
@admin_required
def export_categories_excel():
    """Exporte la liste des catégories en Excel (XLSX) avec titre, en-têtes et style"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    categories = Categorie.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catégories"

    # Styles
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill(start_color="0B2B4A", end_color="0B2B4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Titre
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value="MDL BUSINESS - LISTE DES CATÉGORIES")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Date
    ws.merge_cells('A2:G2')
    date_cell = ws.cell(row=2, column=1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 10

    # En-têtes
    headers = ['ID', 'Nom', 'Slug', 'Description', 'Nombre de produits', 'Statut', 'Date de création']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Données
    row_num = 5
    for cat in categories:
        ws.cell(row=row_num, column=1, value=cat.id)
        ws.cell(row=row_num, column=2, value=cat.nom)
        ws.cell(row=row_num, column=3, value=cat.slug)
        ws.cell(row=row_num, column=4, value=cat.description or '')
        ws.cell(row=row_num, column=5, value=len(cat.produits))
        ws.cell(row=row_num, column=6, value='Actif' if cat.est_active else 'Inactif')
        ws.cell(row=row_num, column=7, value=cat.date_creation.strftime('%d/%m/%Y %H:%M') if cat.date_creation else '')

        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

        row_num += 1

    # Ajuster les largeurs
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20

    # Footer
    ws.merge_cells(f'A{row_num + 1}:G{row_num + 1}')
    footer_cell = ws.cell(row=row_num + 1, column=1, value="MDL Business - Tous droits réservés")
    footer_cell.font = Font(size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = current_app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers.set('Content-Disposition',
                         f'attachment; filename=categories_{datetime.now().strftime("%Y%m%d")}.xlsx')
    response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return response